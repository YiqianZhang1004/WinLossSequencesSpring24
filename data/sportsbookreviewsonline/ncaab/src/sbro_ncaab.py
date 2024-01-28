import openpyxl
from selenium import webdriver  
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
import csv

def check_valid(row1, row2):
    # checks if the two rows have the same date
    if (row1[0].text!=row2[0].text):
        return False
    # checks if the game is not at neutral grounds
    if (row1[2].text != "V" and row1[2].text != "H"):
        return False
    if (row2[2].text != "V" and row2[2].text != "H"):
        return False
    try:
        str(row1[3].text)
        str(row2[3].text)

        int(row1[6].text)
        int(row2[6].text)
    except:
        return False
    return True

def check_valid_xlsx(i):
    # checks if the two rows have the same date
    if (sheet.cell(row=i,column=1).value != sheet.cell(row=i+1,column=1).value):
        return False
    # checks if the game is not at neutral grounds
    if (sheet.cell(row=i,column=3).value != "V" and sheet.cell(row=i,column=3).value!= "H"):
        return False
    if (sheet.cell(row=i+1,column=3).value != "V" and sheet.cell(row=i+1,column=3).value!= "H"):
        return False
    try:
        str(row1[3].text)
        str(row2[3].text)

        int(row1[8].text)
        int(row2[8].text)
    except:
        return False
    return True

final_data = [["date", "season", "team1", "team2", "score1", "score2", "open1", "open2","close1","close2", "moneyline1", "moneyline2","2H1","2H2"]]
seasons = [2007, 2008, 2009, 2010, 2011, 2012, 2013, 2014, 2015, 2016, 2017, 2018,2019,2020]

for season in seasons:
        
    file_path = 'data/sportsbookreviewsonline/ncaab/raw_data/sbro_ncaab_' + str(season) + '.xlsx'

    workbook = openpyxl.load_workbook(file_path)

    sheet = workbook.active

    num_rows = sheet.max_row
    num_cols = sheet.max_column

    for i in range(2, num_rows+1, 2):
        if check_valid_xlsx(i):
            original_date = str(sheet.cell(row=i, column=1).value)
            if len(str(original_date)) == 3:
                month = "0" + original_date[0]
                day = original_date[1:3]
            else:
                month = original_date[0:2]
                day = original_date[2:4]
            processed_date=str(season) + "-" + month + "-" + day

            team1 = str(sheet.cell(row=i, column = 4).value).lower()
            team2 = str(sheet.cell(row=i+1, column=4).value).lower()

            score1 = int(sheet.cell(row=i, column=7).value)
            score2 = int(sheet.cell(row=i+1, column=7).value)

            try:
                open1 = float(sheet.cell(row=i, column=8).value)
                open2 = float(sheet.cell(row=i+1, column=8).value)
            except:
                open1 = "NaN"
                open2 = "NaN"
            
            try:
                close1 = float(sheet.cell(row=i, column=9).value)
                close2 = float(sheet.cell(row=i+1, column=9).value)
            except:
                close1= "NaN"
                close2 = "NaN"

            try:
                moneyline1 = float(sheet.cell(row=i, column=10).value)
                moneyline2 = float(sheet.cell(row=i+1, column=10).value)
            except:
                moneyline1 = "NaN"
                moneyline2 = "NaN"

            try:
                twoH1 = float(sheet.cell(row=i, column=11).value)
                twoH2 = float(sheet.cell(row=i+1, column=11).value)
            except:
                twoH1 = "NaN"
                twoH2 = "NaN"

            if sheet.cell(row=i,column=3) == "V":
                game_data = [processed_date, season, team2, team1, score2, score1, open2, open1, close2, close1, moneyline2, moneyline1, twoH2, twoH1]
            else:
                game_data = [processed_date, season, team1, team2, score1, score2, open1, open2, close1, close2, moneyline1, moneyline2, twoH1, twoH2]
            final_data.append(game_data)


options = Options()
# must be kept false
options.headless = False

driver = webdriver.Chrome(options = options)
driver.get("https://www.sportsbookreviewsonline.com/scoresoddsarchives/ncaa-basketball-2021-22/") 
driver.implicitly_wait(5)

table = driver.find_element(By.CSS_SELECTOR, "table")
rows = table.find_elements(By.CSS_SELECTOR, "tr")

# loops through two rows at a time
for i in range(1,len(rows),2):
    
    row1 = rows[i].find_elements(By.CSS_SELECTOR, "td")
    row2 = rows[i+1].find_elements(By.CSS_SELECTOR, "td")
    if check_valid(row1, row2):
        # getting the date in the right format
        original_date = row1[0].text
        if len(str(original_date)) == 3:
            month = "0" + original_date[0]
            day = original_date[1:3]
        else:
            month = original_date[0:2]
            day = original_date[2:4]
        processed_date=str(season) + "-" + month + "-" + day

        team1 = str(row1[3].text).lower()
        team2 = str(row2[3].text).lower()

        score1 = int(row1[6].text)
        score2 = int(row2[6].text)

        try:
            open1 = float(row1[7].text)
            open2 = float(row2[7].text)
        except:
            open1 = "NaN"
            open2 = "NaN"
        
        try:
            close1 = float(row1[8].text)
            close2 = float(row2[8].text)
        except:
            close1= "NaN"
            close2 = "NaN"

        try:
            moneyline1 = float(row1[9].text)
            moneyline2 = float(row2[9].text)
        except:
            moneyline1 = "NaN"
            moneyline2 = "NaN"

        try:
            twoH1 = float(row1[10].text)
            twoH2 = float(row2[10].text)
        except:
            twoH1 = "NaN"
            twoH2 = "NaN"

        # switch order of data so that home team is always first
        if row1[2].text == "V":
            game_data = [processed_date, season, team2, team1, score2, score1, open2, open1, close2, close1, moneyline2, moneyline1, twoH2, twoH1]
        else:
            game_data = [processed_date, season, team1, team2, score1, score2, open1, open2, close1, close2, moneyline1, moneyline2, twoH1, twoH2]
        final_data.append(game_data)


with open("data/sportsbookreviewsonline/ncaab/processed_data/sbro_ncaab.csv", 'w', newline='') as file:
    csv_writer = csv.writer(file)

    csv_writer.writerows(final_data)